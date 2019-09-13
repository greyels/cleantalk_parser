import openpyxl
from datetime import datetime
from input import URLS, OUTPUT_FILE
from blacklist import BlackList

CT = 'Cleantalk'

def get_bl_emails():
    bl_emails = set()
    for url in URLS:
        black_list = BlackList(url)
        black_list.collect_emails()
        try:
            bl_emails.update(black_list.emails)
        except TypeError:
            pass
    return bl_emails

def get_date_and_time():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
def write_to_file(exec_time, bl_emails, file):
    result = 0
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb[CT]
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.create_sheet(CT, 0)
        ws.cell(column=1, row=1).value = 'Date'
    ws.cell(column=ws.max_column+1, row=1).value = exec_time
    row = 2
    if bl_emails:
        print('The next emails are in the black list:')
        for email in bl_emails:
            print(email)
            ws.cell(column=ws.max_column, row=row).value = email
            row += 1
    else:
        ws.cell(column=ws.max_column, row=row).value = 'No emails!'
    try:
        wb.save(file)
    except PermissionError:
        print('File ' + file + ' is opened.\nCan\'t write to the file.')
        result = 1
    return result

def main():
    exec_time = get_date_and_time()
    bl_emails = get_bl_emails()
    result = write_to_file(exec_time, bl_emails, OUTPUT_FILE)
    return result

if __name__ == '__main__':
    exit(main())
